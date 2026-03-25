# backend/export_utils.py
"""
数据导出工具模块。

提供两类导出：
  1. Excel（openpyxl）：分析汇总、批量结果明细、历史记录
  2. PDF（reportlab）：学术风格实验对比报告
"""
import io
import math
from datetime import datetime

# ──────────────────────────────────────────────
# Excel 导出
# ──────────────────────────────────────────────

def export_dashboard_excel(algo_stats: list, recent_records: list, score_trend: list) -> bytes:
    """
    导出数据大屏汇总 Excel。
    包含三个 Sheet：
      - 算法汇总统计
      - 最近分析记录
      - 评分趋势
    """
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, numbers)
    from openpyxl.utils import get_column_letter

    wb  = Workbook()

    # ── 通用样式 ──
    HEADER_FILL  = PatternFill("solid", fgColor="4F6EF7")
    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
    TITLE_FONT   = Font(bold=True, size=14, color="1A1F36")
    ALT_FILL     = PatternFill("solid", fgColor="F5F7FA")
    GOOD_FILL    = PatternFill("solid", fgColor="E8F5E9")
    BORDER_SIDE  = Side(style='thin', color='E4E7ED')
    THIN_BORDER  = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                          top=BORDER_SIDE, bottom=BORDER_SIDE)
    CENTER  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT    = Alignment(horizontal='left',   vertical='center')
    RIGHT   = Alignment(horizontal='right',  vertical='center')

    def style_header_row(ws, row_num, col_count):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill   = HEADER_FILL
            cell.font   = HEADER_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER

    def style_data_row(ws, row_num, col_count, alt=False):
        fill = ALT_FILL if alt else PatternFill()
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill   = fill
            cell.border = THIN_BORDER
            cell.alignment = CENTER

    def auto_width(ws, min_w=8, max_w=30):
        for col in ws.columns:
            length = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = \
                max(min_w, min(length + 2, max_w))

    # ════════════════════════════════
    # Sheet 1：算法汇总统计
    # ════════════════════════════════
    ws1 = wb.active
    ws1.title = "算法汇总统计"

    # 标题
    ws1.merge_cells('A1:H1')
    ws1['A1'] = f"图像加密算法安全性统计汇总（导出时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}）"
    ws1['A1'].font      = TITLE_FONT
    ws1['A1'].alignment = CENTER
    ws1.row_dimensions[1].height = 30

    headers1 = ['算法名称', '分析次数', '平均安全评分', '平均信息熵',
                 '平均NPCR(%)', '平均UACI(%)', '平均PSNR(dB)', '综合排名']
    for c, h in enumerate(headers1, 1):
        ws1.cell(row=2, column=c, value=h)
    style_header_row(ws1, 2, len(headers1))
    ws1.row_dimensions[2].height = 22

    for i, s in enumerate(algo_stats):
        r = i + 3
        row_data = [
            s.get('algorithm', ''),
            s.get('count', 0),
            s.get('avg_score', 0),
            s.get('avg_entropy', 0),
            s.get('avg_npcr', 0),
            s.get('avg_uaci', 0),
            s.get('avg_psnr', 0),
            i + 1,
        ]
        for c, v in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=v)
            cell.border    = THIN_BORDER
            cell.alignment = LEFT if c == 1 else CENTER
        style_data_row(ws1, r, len(headers1), alt=(i % 2 == 1))
        # 评分高的行标绿
        if s.get('avg_score', 0) >= 80:
            for c in range(1, len(headers1) + 1):
                ws1.cell(row=r, column=c).fill = GOOD_FILL

    auto_width(ws1)

    # ════════════════════════════════
    # Sheet 2：最近分析记录
    # ════════════════════════════════
    ws2 = wb.create_sheet("最近分析记录")
    ws2.merge_cells('A1:G1')
    ws2['A1'] = f"最近分析记录（共 {len(recent_records)} 条）"
    ws2['A1'].font      = TITLE_FONT
    ws2['A1'].alignment = CENTER
    ws2.row_dimensions[1].height = 28

    headers2 = ['文件名', '算法', '安全评分', '信息熵', 'NPCR(%)', 'UACI(%)', '分析时间']
    for c, h in enumerate(headers2, 1):
        ws2.cell(row=2, column=c, value=h)
    style_header_row(ws2, 2, len(headers2))

    for i, rec in enumerate(recent_records):
        r = i + 3
        m = rec.get('metrics', {}) or {}
        row_data = [
            rec.get('original_filename', ''),
            rec.get('algorithm', ''),
            rec.get('security_score', 0),
            round(m.get('entropy_encrypted', 0) or 0, 4),
            round(m.get('npcr', 0) or 0, 3),
            round(m.get('uaci', 0) or 0, 3),
            str(rec.get('created_at', ''))[:16],
        ]
        for c, v in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=v)
            cell.border    = THIN_BORDER
            cell.alignment = LEFT if c in (1, 2, 7) else CENTER
        style_data_row(ws2, r, len(headers2), alt=(i % 2 == 1))

    auto_width(ws2)

    # ════════════════════════════════
    # Sheet 3：评分趋势
    # ════════════════════════════════
    ws3 = wb.create_sheet("评分趋势")
    ws3.merge_cells('A1:D1')
    ws3['A1'] = "安全评分历史趋势"
    ws3['A1'].font      = TITLE_FONT
    ws3['A1'].alignment = CENTER
    ws3.row_dimensions[1].height = 28

    headers3 = ['序号', '算法', '安全评分', '分析时间']
    for c, h in enumerate(headers3, 1):
        ws3.cell(row=2, column=c, value=h)
    style_header_row(ws3, 2, len(headers3))

    for i, rec in enumerate(score_trend):
        r = i + 3
        row_data = [i + 1, rec.get('algorithm', ''), rec.get('score', 0), rec.get('created_at', '')]
        for c, v in enumerate(row_data, 1):
            cell = ws3.cell(row=r, column=c, value=v)
            cell.border    = THIN_BORDER
            cell.alignment = CENTER
        style_data_row(ws3, r, len(headers3), alt=(i % 2 == 1))

    auto_width(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_batch_excel(batch_result: dict) -> bytes:
    """
    导出单次批量处理结果 Excel。
    包含两个 Sheet：
      - 算法评分汇总
      - 逐条明细
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    HEADER_FILL = PatternFill("solid", fgColor="4F6EF7")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    TITLE_FONT  = Font(bold=True, size=13, color="1A1F36")
    ALT_FILL    = PatternFill("solid", fgColor="F5F7FA")
    GOOD_FILL   = PatternFill("solid", fgColor="E8F5E9")
    FAIL_FILL   = PatternFill("solid", fgColor="FEF0F0")
    BORDER_SIDE = Side(style='thin', color='E4E7ED')
    THIN_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                         top=BORDER_SIDE, bottom=BORDER_SIDE)
    CENTER = Alignment(horizontal='center', vertical='center')
    LEFT   = Alignment(horizontal='left',   vertical='center')

    def set_header(ws, row, headers):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.alignment = CENTER
            cell.border    = THIN_BORDER
        ws.row_dimensions[row].height = 22

    def auto_width(ws, min_w=8, max_w=28):
        for col in ws.columns:
            length = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = \
                max(min_w, min(length + 2, max_w))

    summary = batch_result.get('summary', [])
    results = batch_result.get('results', [])
    total   = batch_result.get('total', 0)
    success = batch_result.get('success', 0)
    failed  = batch_result.get('failed', 0)

    # ── Sheet 1：汇总 ──
    ws1 = wb.active
    ws1.title = "算法评分汇总"
    ws1.merge_cells('A1:G1')
    ws1['A1'] = (f"批量处理结果汇总  |  总任务 {total}  成功 {success}  失败 {failed}"
                 f"  |  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws1['A1'].font      = TITLE_FONT
    ws1['A1'].alignment = CENTER
    ws1.row_dimensions[1].height = 28

    h1 = ['算法', '图片数', '平均安全评分', '平均耗时(ms)', '平均信息熵', '平均NPCR(%)', '平均UACI(%)']
    set_header(ws1, 2, h1)

    for i, s in enumerate(summary):
        r = i + 3
        data = [s.get('algorithm',''), s.get('count',0), s.get('avg_score',0),
                s.get('avg_time_ms',0), s.get('avg_entropy',0),
                s.get('avg_npcr',0), s.get('avg_uaci',0)]
        for c, v in enumerate(data, 1):
            cell = ws1.cell(row=r, column=c, value=v)
            cell.border    = THIN_BORDER
            cell.alignment = LEFT if c == 1 else CENTER
            cell.fill      = ALT_FILL if i % 2 == 1 else PatternFill()
        if s.get('avg_score', 0) >= 80:
            for c in range(1, len(h1)+1):
                ws1.cell(row=r, column=c).fill = GOOD_FILL

    auto_width(ws1)

    # ── Sheet 2：逐条明细 ──
    ws2 = wb.create_sheet("逐条明细")
    ws2.merge_cells('A1:I1')
    ws2['A1'] = f"批量处理逐条明细（共 {len(results)} 条）"
    ws2['A1'].font      = TITLE_FONT
    ws2['A1'].alignment = CENTER
    ws2.row_dimensions[1].height = 26

    h2 = ['文件名', '算法', '安全评分', '耗时(ms)', '信息熵',
          'NPCR(%)', 'UACI(%)', 'PSNR(dB)', '状态']
    set_header(ws2, 2, h2)

    for i, rec in enumerate(results):
        r = i + 3
        m = rec.get('metrics', {}) or {}
        ok = rec.get('success', False)
        data = [
            rec.get('original_filename', ''),
            rec.get('algorithm', ''),
            rec.get('security_score', 0) if ok else '-',
            rec.get('encrypt_time_ms', 0) if ok else '-',
            round(m.get('entropy_encrypted', 0) or 0, 4) if ok else '-',
            round(m.get('npcr', 0) or 0, 3)             if ok else '-',
            round(m.get('uaci', 0) or 0, 3)             if ok else '-',
            round(m.get('psnr', 0) or 0, 2)             if ok else '-',
            '✓ 成功' if ok else f'✗ {rec.get("error","失败")}',
        ]
        for c, v in enumerate(data, 1):
            cell = ws2.cell(row=r, column=c, value=v)
            cell.border    = THIN_BORDER
            cell.alignment = LEFT if c in (1, 2, 9) else CENTER
        fill = GOOD_FILL if ok else FAIL_FILL
        for c in range(1, len(h2)+1):
            ws2.cell(row=r, column=c).fill = fill if (not ok or i%2==0) else ALT_FILL

    auto_width(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ──────────────────────────────────────────────
# PDF 导出（实验对比报告）
# ──────────────────────────────────────────────

def export_experiment_pdf(experiment_data: list, title: str = "图像加密算法实验对比报告") -> bytes:
    """
    导出实验对比 PDF 报告（学术风格）。
    包含：
      - 封面（标题、时间、数据规模）
      - 实验指标说明
      - 算法对比表格（均值 ± 标准差）
      - 各算法说明
      - 结论与分析
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                     Table, TableStyle, HRFlowable, PageBreak)
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
    import os

    # ── 注册中文字体 ──
    # 尝试常见 macOS/Linux 中文字体路径
    _font_registered = False
    font_candidates = [
        '/System/Library/Fonts/PingFang.ttc',
        '/System/Library/Fonts/STHeiti Medium.ttc',
        '/Library/Fonts/Arial Unicode MS.ttf',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        '/usr/share/fonts/wqy-microhei/wqy-microhei.ttc',
    ]
    for fp in font_candidates:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont('CJK', fp))
                _font_registered = True
                break
            except Exception:
                continue

    CN_FONT  = 'CJK' if _font_registered else 'Helvetica'
    CN_BOLD  = 'CJK' if _font_registered else 'Helvetica-Bold'

    # ── 颜色定义 ──
    PRIMARY   = colors.HexColor('#4F6EF7')
    HEADER_BG = colors.HexColor('#4F6EF7')
    ALT_BG    = colors.HexColor('#F5F7FA')
    BEST_BG   = colors.HexColor('#E8F5E9')
    BORDER_C  = colors.HexColor('#E4E7ED')
    TEXT_DARK = colors.HexColor('#1A1F36')
    TEXT_GRAY = colors.HexColor('#606266')

    # ── 样式 ──
    styles  = getSampleStyleSheet()
    s_title = ParagraphStyle('Title2', fontName=CN_BOLD,  fontSize=20, textColor=TEXT_DARK,
                              alignment=TA_CENTER, spaceAfter=6, leading=28)
    s_sub   = ParagraphStyle('Sub',   fontName=CN_FONT,  fontSize=12, textColor=TEXT_GRAY,
                              alignment=TA_CENTER, spaceAfter=4)
    s_h2    = ParagraphStyle('H2',    fontName=CN_BOLD,  fontSize=13, textColor=PRIMARY,
                              spaceBefore=14, spaceAfter=6, leading=18)
    s_body  = ParagraphStyle('Body',  fontName=CN_FONT,  fontSize=10, textColor=TEXT_DARK,
                              leading=16, spaceAfter=4, alignment=TA_JUSTIFY)
    s_note  = ParagraphStyle('Note',  fontName=CN_FONT,  fontSize=8,  textColor=TEXT_GRAY,
                              leading=12, spaceAfter=4)
    s_cell  = ParagraphStyle('Cell',  fontName=CN_FONT,  fontSize=8,  textColor=TEXT_DARK,
                              alignment=TA_CENTER, leading=11)
    s_cell_bold = ParagraphStyle('CellB', fontName=CN_BOLD, fontSize=8, textColor=TEXT_DARK,
                                  alignment=TA_CENTER, leading=11)
    s_cell_l = ParagraphStyle('CellL', fontName=CN_FONT, fontSize=8, textColor=TEXT_DARK,
                               alignment=TA_LEFT, leading=11)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2.5*cm, bottomMargin=2*cm,
        title=title,
        author="图像加密分析平台",
    )

    story = []
    W = A4[0] - 4*cm   # 可用宽度

    def hr(): return HRFlowable(width='100%', thickness=1, color=BORDER_C, spaceAfter=8)

    # ════════════════════════════════
    # 封面
    # ════════════════════════════════
    story.append(Spacer(1, 2*cm))
    story.append(Paragraph(title, s_title))
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph("Digital Image Encryption Algorithm Experimental Comparison Report", s_sub))
    story.append(Spacer(1, 0.5*cm))
    story.append(hr())

    total_samples = sum(r.get('count', 0) for r in experiment_data)
    algo_count    = len(experiment_data)
    story.append(Paragraph(
        f"生成时间：{datetime.now().strftime('%Y年%m月%d日 %H:%M')}　|　"
        f"参与对比算法：{algo_count} 种　|　总样本量：{total_samples} 条",
        s_sub
    ))
    story.append(Spacer(1, 1*cm))

    # 摘要
    story.append(Paragraph("摘　要", s_h2))
    story.append(Paragraph(
        "本报告基于图像加密分析平台的历史实验数据，对多种图像加密算法进行综合性能评估。"
        "评估指标涵盖信息熵、像素相关性、直方图均匀性（卡方检验）、差分攻击（NPCR/UACI）、"
        "峰值信噪比（PSNR）、平均运行长度（ARL）以及加密耗时。"
        "所有数值均以均值 ± 标准差的形式呈现，反映算法在不同图像上的稳定性。",
        s_body
    ))

    story.append(PageBreak())

    # ════════════════════════════════
    # 实验指标说明
    # ════════════════════════════════
    story.append(Paragraph("一、评估指标说明", s_h2))

    metrics_info = [
        ("信息熵（Entropy）",    "衡量密文图像的随机性。理想值接近 8.0 bit，优秀阈值 > 7.9。"),
        ("NPCR（像素变化率）",   "原始图像与加密图像像素不同的比例。优秀阈值 > 99.6%，反映明文敏感性。"),
        ("UACI（平均变化强度）", "加密前后像素值差异的平均强度。理想值约 33.46%。"),
        ("PSNR（峰值信噪比）",   "越低表示加密后图像与原图差异越大，加密效果越好。"),
        ("ARL（平均运行长度）",  "衡量图像随机性，越小越好。优秀阈值 < 2.0。"),
        ("卡方检验（Chi²）",     "直方图均匀性度量，值越小说明像素分布越均匀。优秀阈值 < 1000。"),
        ("加密耗时（ms）",       "完成一次图像加密所需时间，反映算法的计算效率。"),
    ]
    for name, desc in metrics_info:
        story.append(Paragraph(f"<b>{name}</b>：{desc}", s_body))

    story.append(Spacer(1, 0.5*cm))

    # ════════════════════════════════
    # 算法对比表格
    # ════════════════════════════════
    story.append(Paragraph("二、算法性能对比表", s_h2))
    story.append(Paragraph(
        "表中数据格式为「均值 ± 标准差」，绿色背景表示该列最优值。",
        s_note
    ))
    story.append(Spacer(1, 0.2*cm))

    # 定义表格列（竖排两组，避免过宽）
    # 第一组：安全性指标
    _build_comparison_table(story, experiment_data, W, s_cell, s_cell_bold, s_cell_l,
                             HEADER_BG, ALT_BG, BEST_BG, BORDER_C, colors.white,
                             group='security')

    story.append(Spacer(1, 0.4*cm))
    story.append(Paragraph("（续表）性能指标对比", s_note))
    story.append(Spacer(1, 0.2*cm))

    _build_comparison_table(story, experiment_data, W, s_cell, s_cell_bold, s_cell_l,
                             HEADER_BG, ALT_BG, BEST_BG, BORDER_C, colors.white,
                             group='performance')

    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph(
        "注：↑ 表示越大越好，↓ 表示越小越好，≈ 表示越接近该值越好。"
        "综合安全评分综合了信息熵、相关性、卡方检验等多项指标，满分 100 分。",
        s_note
    ))

    story.append(PageBreak())

    # ════════════════════════════════
    # 结论
    # ════════════════════════════════
    story.append(Paragraph("三、实验结论", s_h2))

    if experiment_data:
        # 找各项最优
        def best_mean(key, higher=True):
            valid = [(r['algorithm'], r['metrics'].get(key, {}).get('mean', 0))
                     for r in experiment_data if r['metrics'].get(key)]
            if not valid: return '-', 0
            return max(valid, key=lambda x: x[1]) if higher else min(valid, key=lambda x: x[1])

        best_score_algo, best_score_val   = best_mean('security_score', True)
        best_entropy_algo, best_ent_val   = best_mean('entropy_encrypted', True)
        best_npcr_algo, best_npcr_val     = best_mean('npcr', True)
        fastest_algo, fastest_val         = best_mean('encrypt_time_ms', False)

        def uaci_closest():
            valid = [(r['algorithm'], r['metrics'].get('uaci', {}).get('mean', 999))
                     for r in experiment_data]
            return min(valid, key=lambda x: abs(x[1] - 33.46))[0]

        story.append(Paragraph(
            f"根据本次实验数据（共 {total_samples} 条分析记录，{algo_count} 种算法），"
            f"得出以下结论：",
            s_body
        ))

        conclusions = [
            f"综合安全评分最高的算法为 <b>{best_score_algo}</b>，"
            f"平均评分 {best_score_val:.1f}/100，表现出良好的整体安全性。",
            f"信息熵最高的算法为 <b>{best_entropy_algo}</b>（均值 {best_ent_val:.4f} bit），"
            f"接近理论最大值 8.0 bit，加密后图像随机性优秀。",
            f"NPCR 最高的算法为 <b>{best_npcr_algo}</b>（均值 {best_npcr_val:.3f}%），"
            f"对明文单比特变化极其敏感。",
            f"UACI 最接近理想值（33.46%）的算法为 <b>{uaci_closest()}</b>，"
            f"具有均匀的像素变化强度。",
            f"加密速度最快的算法为 <b>{fastest_algo}</b>（均值 {fastest_val:.1f} ms），"
            f"适合对实时性要求较高的场景。",
            "在安全性与性能的综合权衡中，推荐优先考虑现代认证加密算法"
            "（AES-GCM、ChaCha20），它们在信息熵、NPCR 等安全指标上表现优秀，"
            "同时具备合理的加密速度和完整性保护能力。",
        ]
        for c in conclusions:
            story.append(Paragraph(f"• {c}", s_body))

    story.append(Spacer(1, 1*cm))
    story.append(hr())
    story.append(Paragraph(
        f"本报告由图像加密分析平台自动生成  |  {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        s_note
    ))

    doc.build(story)
    return buf.getvalue()


def _build_comparison_table(story, experiment_data, W,
                              s_cell, s_cell_bold, s_cell_l,
                              header_bg, alt_bg, best_bg, border_c, white,
                              group='security'):
    """构建一个对比子表格并追加到 story"""
    from reportlab.platypus import Table, TableStyle, Paragraph

    if group == 'security':
        col_defs = [
            ('安全评分↑\n(0-100)',      'security_score',    True,  False, 33.46),
            ('信息熵↑\n(bit)',          'entropy_encrypted', True,  False, None),
            ('NPCR(%)↑',               'npcr',              True,  False, None),
            ('UACI(%)≈33.46',          'uaci',              False, True,  33.46),
            ('卡方检验↓',              'chi2_encrypted',    False, False, None),
        ]
    else:
        col_defs = [
            ('ARL↓',                   'arl_encrypted',     False, False, None),
            ('水平相关性↓',            'correlation_h_encrypted', False, False, None),
            ('PSNR(dB)↓',             'psnr',              False, False, None),
            ('加密耗时(ms)↓',          'encrypt_time_ms',   False, False, None),
        ]

    n_cols   = len(col_defs) + 1   # +1 for algorithm column
    n_rows   = len(experiment_data) + 2  # header + data + count row

    col_w    = W / n_cols
    algo_w   = col_w * 1.6
    data_w   = (W - algo_w) / len(col_defs)

    # 计算每列的最优值（用于标绿）
    best_vals = {}
    for col_name, key, higher, closest, target in col_defs:
        vals = [(r['algorithm'], r['metrics'].get(key, {}).get('mean'))
                for r in experiment_data if r['metrics'].get(key) is not None]
        vals = [(a, v) for a, v in vals if v is not None]
        if not vals: continue
        if closest:
            best_algo = min(vals, key=lambda x: abs(x[1] - target))[0]
        elif higher:
            best_algo = max(vals, key=lambda x: x[1])[0]
        else:
            best_algo = min(vals, key=lambda x: x[1])[0]
        best_vals[key] = best_algo

    def fmt(obj, digits=2):
        if not obj: return '-'
        m = obj.get('mean'); s = obj.get('std', 0)
        if m is None: return '-'
        return f"{m:.{digits}f}±{s:.{digits}f}"

    def digits_for(key):
        d = {'security_score': 1, 'entropy_encrypted': 4,
             'npcr': 3, 'uaci': 3, 'chi2_encrypted': 0,
             'arl_encrypted': 3, 'correlation_h_encrypted': 4,
             'psnr': 2, 'encrypt_time_ms': 1}
        return d.get(key, 2)

    # 构建表格数据
    header_row = [Paragraph('算法', s_cell_bold)]
    for col_name, _, _, _, _ in col_defs:
        header_row.append(Paragraph(col_name, s_cell_bold))

    count_row = [Paragraph('样本量 n', s_cell)]
    for _, key, _, _, _ in col_defs:
        count_row.append(Paragraph('', s_cell))

    table_data = [header_row]
    cell_styles = []   # 用于标绿

    for row_idx, rec in enumerate(experiment_data):
        row = [Paragraph(rec['algorithm'], s_cell_l)]
        for col_idx, (_, key, _, _, _) in enumerate(col_defs):
            val   = fmt(rec['metrics'].get(key, {}), digits_for(key))
            is_b  = best_vals.get(key) == rec['algorithm']
            style = s_cell_bold if is_b else s_cell
            row.append(Paragraph(val, style))
            if is_b:
                # row+1 因为 header 占第 0 行
                cell_styles.append(('BACKGROUND', (col_idx+1, row_idx+1),
                                     (col_idx+1, row_idx+1), best_bg))
        table_data.append(row)

    col_widths = [algo_w] + [data_w] * len(col_defs)
    t = Table(table_data, colWidths=col_widths, repeatRows=1)

    base_style = [
        ('BACKGROUND',  (0, 0), (-1, 0),  header_bg),
        ('TEXTCOLOR',   (0, 0), (-1, 0),  white),
        ('GRID',        (0, 0), (-1, -1), 0.5, border_c),
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [white, alt_bg]),
        ('TOPPADDING',  (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING',(0,0),(-1,-1), 4),
    ]
    t.setStyle(TableStyle(base_style + cell_styles))
    story.append(t)
